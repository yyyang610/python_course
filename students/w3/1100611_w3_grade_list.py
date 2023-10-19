import openpyxl

#oop
class grade_list:
    
    def __init__(self):
        excel_file_path = 'C:\\Users\Girl\Desktop\grade_list.xlsx'
        workbook = openpyxl.load_workbook(excel_file_path)
        # Get the sheet names
        self.sheet_names=workbook.sheetnames
        self.sheet_name=self.sheet_names[0]
        # Print the sheet names
        print("Sheet names:", self.sheet_names)
        self.sheet = workbook[self.sheet_name]      #sheet name
        self.cell_range_pattern='c5:j12'        #讀取資料的範圍存入變數
        self.cell_values_2d=[]                  #設置空的二維陣列讓下面func都可以用
        
        self.keyin()        #呼叫keyin 把從excel讀出來的資料放入二維陣列
        self.sum_aver_rank()    #排名、總分、平均計算好放入二維陣列
        self.show()     #print出數據
        workbook.close()
    
    #把從excel讀出來的資料放入二維陣列
    def keyin(self):
        for row in self.sheet[self.cell_range_pattern]:
            row_values = []  # Create a list for each row
            for cell in row:
                    row_values.append(cell.value)
            self.cell_values_2d.append(row_values)

    #計算總分、平均、排名
    def sum_aver_rank(self):
        row_len=len(self.cell_values_2d)
        for i in range(1,row_len):
                self.cell_values_2d[i][5]=self.cell_values_2d[i][1]+self.cell_values_2d[i][2]+self.cell_values_2d[i][3]+self.cell_values_2d[i][4]
                self.cell_values_2d[i][6]=self.cell_values_2d[i][5]/4
                
        grade_list_sort = sorted(self.cell_values_2d[1:], key=lambda x: x[5])
        for i in range(len(grade_list_sort)):    
            for j in range(1, len(grade_list_sort[i])):
                grade_list_sort[i][7]=i+1

    #print
    def show(self):
        print("2D Cell values:")
        for row_values in self.cell_values_2d:
            print(row_values)
           
s1=grade_list()


#pop
excel_file_path = 'C:\\Users\Girl\Desktop\grade_list.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)
        # Get the sheet names
sheet_names=workbook.sheetnames
sheet_name=sheet_names[0]
        # Print the sheet names
print("Sheet names:", sheet_names)
sheet = workbook[sheet_name]      #sheet name
cell_range_pattern='c5:j12'        #讀取資料的範圍存入變數
cell_values_2d=[]                  #設置空的二維陣列讓下面func都可以用
        
for row in sheet[cell_range_pattern]:
    row_values = []  # Create a list for each row
    for cell in row:
            row_values.append(cell.value)
    cell_values_2d.append(row_values)

row_len=len(cell_values_2d)
for i in range(1,row_len):
    cell_values_2d[i][5]=cell_values_2d[i][1]+cell_values_2d[i][2]+cell_values_2d[i][3]+cell_values_2d[i][4]
    cell_values_2d[i][6]=cell_values_2d[i][5]/4
                
grade_list_sort = sorted(cell_values_2d[1:], key=lambda x: x[5])
for i in range(len(grade_list_sort)):    
    for j in range(1, len(grade_list_sort[i])):
        grade_list_sort[i][7]=i+1

print("2D Cell values:")
for row_values in cell_values_2d:
    print(row_values)

workbook.close()