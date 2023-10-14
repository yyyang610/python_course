
import openpyxl

#POP
excel_file_path = 'C:/Users/julia/OneDrive/桌面/python_course/teacher/w3/grade_list.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)


sheet_names = workbook.sheetnames
sheet_name=sheet_names[0]
print("Sheet names:", sheet_names)
sheet = workbook[sheet_name]
print(sheet)

cell_range_pattern='c5:j12'
cell_range = sheet[cell_range_pattern]

cell_values_2d = [] # Initialize a 2D list to store cell values

# Iterate through the cells in the specified range and collect values
for row in sheet[cell_range_pattern]:
    row_values = []  # for each row
    for cell in row:
        row_values.append(cell.value)
    cell_values_2d.append(row_values)  # Append the row list

workbook.close()

for i in range(1,len(cell_values_2d)):
    sum=0
    for j in range(1,5):
        sum+=cell_values_2d[i][j]
    cell_values_2d[i][5]=sum
    cell_values_2d[i][6]=sum/4

A=[]
for i in range(1,len(cell_values_2d)):
    A.append(cell_values_2d[i][5])
A.sort()

for i in range(1,len(cell_values_2d)):
    cell_values_2d[i][7]=len(cell_values_2d)-A.index(cell_values_2d[i][5])-1

for i in range(len(cell_values_2d)):
    for j in range(8):
        print(cell_values_2d[i][j],end='\t')
    print() #印出來



#OOP
class Grade:
    def __init__(self,excel_file_path):
        self.excel_file_path=excel_file_path
        self.workbook = openpyxl.load_workbook(excel_file_path)

        self.cell_range_pattern='c5:j12'
        self.cell_range = sheet[cell_range_pattern]
        self.cell_values_2d = [] # Initialize a 2D list to store cell values
        self.sheetName()
        self.collectValue()
        self.sumAvg()
        self.rank()
        self.printt()
        

    def sheetName(self):
        sheet_names = self.workbook.sheetnames
        sheet_name=sheet_names[0]
        print("Sheet names:", sheet_names)
        self.sheet = self.workbook[sheet_name]

    def collectValue(self):
        self.cell_range = self.sheet[self.cell_range_pattern]
        # Iterate through the cells in the specified range and collect values
        for row in self.sheet[self.cell_range_pattern]:
            row_values = []  # for each row
            for cell in row:
                row_values.append(cell.value)
            self.cell_values_2d.append(row_values)  # Append the row list
        workbook.close()
        

    #sum and average
    def sumAvg(self):
        for i in range(1,len(self.cell_values_2d)):
            sum=0
            for j in range(1,5):
                sum+=self.cell_values_2d[i][j]
                self.cell_values_2d[i][5]=sum
                self.cell_values_2d[i][6]=sum/4
    #rank
    def rank(self):
        A=[]
        for i in range(1,len(self.cell_values_2d)):
            A.append(self.cell_values_2d[i][5])
        A.sort()

        for i in range(1,len(self.cell_values_2d)):
                self.cell_values_2d[i][7]=len(self.cell_values_2d)-A.index(self.cell_values_2d[i][5])-1
    #print()
    def printt(self):
         for i in range(len(self.cell_values_2d)):
            for j in range(8):
                 print(self.cell_values_2d[i][j],end='\t')
            print()
            
oop=Grade('C:/Users/julia/OneDrive/桌面/python_course/teacher/w3/grade_list.xlsx')