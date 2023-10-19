
import openpyxl
# Replace with the path to your Excel file
excel_file_path = 'C:\\Users\\李芷榆\\OneDrive\\桌面\\python\\grade_list.xlsx'

# Load the Excel workbook and select the sheet
workbook = openpyxl.load_workbook(excel_file_path)
# Get the sheet names
sheet_names = workbook.sheetnames
sheet_name=sheet_names[0]
# Print the sheet names
print("Sheet names:", sheet_names)
sheet = workbook[sheet_name]

cell_range_pattern='c5:j12'
cell_range = sheet[cell_range_pattern]

#pop
a=[]
grade_table=[]
for i in cell_range:
    line=[]
    for j in i:
        line.append(j.value)
    grade_table.append(line)
workbook.close()

for i in range(1,8):
    grade_table[i][5]=0
    grade_table[i][6]=0
    grade_table[i][7]=0
    for j in range(1,5):
        grade_table[i][5]+=grade_table[i][j]
    a.append(grade_table[i][5])
    grade_table[i][6]=grade_table[i][5]/4

a.sort(reverse=True)
for i in range(1,8):
    for j in range(1,5):
        grade_table[i][7]=a.index(grade_table[i][5])+1

print(grade_table)

#oop
class grade_table:
    def __init__(self,import_excel):
        self.import_excel=import_excel
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet_names = workbook.sheetnames
        sheet_name=sheet_names[0]
        sheet = workbook[sheet_name]
        cell_range_pattern='c5:j12'
        cell_range = sheet[cell_range_pattern]
        
        
        self.grade_table=[]
        for i in cell_range:
            line=[]
            for j in i:
                line.append(j.value)
            self.grade_table.append(line)
        workbook.close()
    def calculation(self):
        a=[]
        for i in range(1,8):
            self.grade_table[i][5]=0
            self.grade_table[i][6]=0
            self.grade_table[i][7]=0
            for j in range(1,5):
                self.grade_table[i][5]+=self.grade_table[i][j]
            a.append(self.grade_table[i][5])
            self.grade_table[i][6]=self.grade_table[i][5]/4
        a.sort(reverse=True)
        for i in range(1,8):
            for j in range(1,5):
                self.grade_table[i][7]=a.index(self.grade_table[i][5])+1        
    def print(self):
        print(self.grade_table)

mygrade=grade_table('C:\\Users\\李芷榆\\OneDrive\\桌面\\python\\grade_list.xlsx')
mygrade.calculation()
mygrade.print()


