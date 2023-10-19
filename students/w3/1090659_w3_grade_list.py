import sys;   
import datetime
import openpyxl

print("POP")
excel_file_path = "C:\Users\max\Desktop\python\python_course\teacher\w3\grade_list.xlsx"
workbook=openpyxl.load_workbook(excel_file_path)

sheet_names = workbook.sheetnames
sheet_name = workbook[sheet_names[0]]
print("Sheet names:", sheet_names)
sheet=workbook[sheet_name]

cell_range_pattern = 'c5:j12'
grade_list = [[data.value for data in row] for row in cell_range_pattern]
cell_range = sheet[cell_range_pattern]
print(cell_range)
workbook.close()

for i in range(len(grade_list)):
    grade_list[i][5]=grade_list[i][1]+grade_list[i][2]+grade_list[i][3]+grade_list[i][4]
    grade_list[i][6]=grade_list[i][5]/4
    
sort_data = sorted(grade_list, key=lambda x: x[6])

for i in range(len(sort_data)):
    for j in range(len(grade_list)):
        if(sort_data[i][6]==grade_list[j][6]):
            grade_list[j][7]=i+1


for i in range(len(grade_list)):
    print(grade_list[i][0], end='\t')
    for j in range(1, len(grade_list[i])):
        print(grade_list[i][j], end='\t')
    print()

print("----")
print("OOP")

class Account:
    def __init__(self):
        excel_file_path = "C:\Users\max\Desktop\python\python_course\teacher\w3\grade_list.xlsx"
        workbook=openpyxl.load_wordbook(excel_file_path)

        self.sheet_names = workbook.sheetnames
        self.sheet_name = workbook[sheet_names[0]]
        print("Sheet names:", sheet_names)
        self.sheet=workbook[sheet_name]

        self.cell_range_pattern = 'c5:j12'
        self.grade_list = [[data.value for data in row] for row in cell_range_pattern]
        self.cell_range = sheet[cell_range_pattern]
        print(cell_range)
        workbook.close()

    def caculate(self):
        for i in range(len(self.grade_list)):
            self.grade_list[i][5] = self.grade_list[i][1] + self.grade_list[i][2] + self.grade_list[i][3] + self.grade_list[i][4]
            self.grade_list[i][6] = self.grade_list[i][5] / 4

    
        sort_data = sorted(grade_list, key=lambda x: x[6])

        for i in range(len(sort_data)):
            for j in range(len(grade_list)):
                if(sort_data[i][6]==grade_list[j][6]):
                    grade_list[j][7]=i+1


        for i in range(len(grade_list)):
            print(grade_list[i][0], end='\t')
            for j in range(1, len(grade_list[i])):
                print(grade_list[i][j], end='\t')
            print()


  