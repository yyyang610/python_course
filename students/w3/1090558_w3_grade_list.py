import openpyxl

#==================== Get data ====================
# Replace with the path to your Excel file
excel_file_path = 'C:/Users/User/Desktop/grade_list.xlsx'
# Load the Excel workbook and select the sheet
workbook = openpyxl.load_workbook(excel_file_path)
# Get the sheet names
sheet_names = workbook.sheetnames
sheet_name=sheet_names[0]
# Print the sheet names
sheet = workbook[sheet_name]
# Specify the cell coordinates (row and column indices, 1-based index)
# Access the cell range
cell_range_pattern='c5:j12'
cell_range = sheet[cell_range_pattern]
# Initialize a 2D list to store cell values
cell_values_2d = []
# Iterate through the cells in the specified range and collect values
for row in sheet[cell_range_pattern]:
    row_values = []  # Create a list for each row
    for cell in row:
        row_values.append(cell.value)
    cell_values_2d.append(row_values)  # Append the row list to the 2D list
workbook.close()


#==================== POP ====================
#score, average
for i in range(1, len(cell_values_2d)):
    score = 0
    for j in range(1, 5):
        score += cell_values_2d[i][j]
    cell_values_2d[i][5] = score
    cell_values_2d[i][6] = score / 4

#rank
for i in range(1, len(cell_values_2d)):
    k = 1 
    for j in range (1, len(cell_values_2d)):
        if(i == j):
            continue
        else:
            if(cell_values_2d[i][5] < cell_values_2d[j][5]):
                k += 1
    cell_values_2d[i][7] = k

#print 
for i in range(len(cell_values_2d)):
    print(cell_values_2d[i][0], end='\t')
    for j in range(1, len(cell_values_2d[i])):
        print(cell_values_2d[i][j], end='\t')
    print()
print()


#==================== OOP ====================
class Grade_list():
    def __init__(self, file_data):
        self.file_data = file_data
        self.workbook = openpyxl.load_workbook(file_data)
        self.cell_range_pattern = 'c5:j12'
        self.data_value_2d = []

        self.get_data()
        self.total_core()
        self.average()
        self.print()
        self.close()

    def get_data(self):
        self.sheet_names = self.workbook.sheetnames
        self.sheet_name = self.sheet_names[0]
        sheet_name = self.sheet_names[0]
        sheet = self.workbook[sheet_name]
        
        for row in sheet[cell_range_pattern]:
            row_values = []  # Create a list for each row
            for cell in row:
                row_values.append(cell.value)
            self.data_value_2d.append(row_values)  # Append the row list to the 2D list
    
    def total_core(self):
        for i in range(1, len(self.data_value_2d)):
            score = 0
            for j in range(1, 5):
                score += self.data_value_2d[i][j]
            self.data_value_2d[i][5] = score
            self.data_value_2d[i][6] = score / 4

    def average(self):
        for i in range(1, len(self.data_value_2d)):
            k = 1 
            for j in range (1, len(self.data_value_2d)):
                if(i == j):
                    continue
                else:
                    if(self.data_value_2d[i][5] < self.data_value_2d[j][5]):
                        k += 1
            self.data_value_2d[i][7] = k


    def print(self):
        for i in range(len(self.data_value_2d)):
            print(self.data_value_2d[i][0], end='\t')
            for j in range(1, len(self.data_value_2d[i])):
                print(self.data_value_2d[i][j], end='\t')
            print()
        print()

    def close(self):
        self.workbook.close()
    
#test OOP
w3_OOP = Grade_list('C:/Users/User/Desktop/grade_list.xlsx')