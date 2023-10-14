import openpyxl
# oop
class Grade:
    def __init__(self, file_path: str, data_range: str):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[workbook.sheetnames[0]]
        call_range = sheet[data_range]
        self.data = list()
        for row in sheet[data_range]:
            row_value = list()
            for cell in row:
                row_value.append(cell.value)
            self.data.append(row_value)
        self.__set_sum_and_average()
        self.__set_ranking()
        # print(self.data)
        workbook.close()
    def print_data(self):
        for row in self.data:
            print(row)
    def __set_sum_and_average(self):
        for row in self.data[1:]:
            sum = 0
            for score in row[1:5]:
                sum += score
            row[5] = sum
            row[6] = sum / 4

    def __set_ranking(self):
        rank_dict = dict()
        for i in range(1,len(self.data)):
            rank_dict[i] = self.data[i][5]
        rank_dict = sorted(rank_dict.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
        for rank in range(len(rank_dict)):
            self.data[rank_dict[rank][0]][7] = rank+1

print("OOP")
oop = Grade("../../teacher/w3/grade_list.xlsx", 'c5:j12')
oop.print_data()

# pop
print("POP")
workbook = openpyxl.load_workbook("../../teacher/w3/grade_list.xlsx")
sheet = workbook[workbook.sheetnames[0]]
call_range = sheet['c5:j12']
data = list()
for row in sheet['c5:j12']:
    row_value = list()
    for cell in row:
        row_value.append(cell.value)
    data.append(row_value)

for row in data[1:]:
    sum = 0
    for score in row[1:5]:
        sum += score
    row[5] = sum
    row[6] = sum / 4

rank_dict = dict()
for i in range(1, len(data)):
    rank_dict[i] = data[i][5]
rank_dict = sorted(rank_dict.items(), key=lambda kv: (kv[1], kv[0]), reverse=True)
for rank in range(len(rank_dict)):
    data[rank_dict[rank][0]][7] = rank + 1

for row in data:
    print(row)
