
# pip install openpyxl
# show read excel 

import openpyxl

# Replace with the path to your Excel file
excel_file_path = 'D:/MyData/Desktop/PYTHON/python_course/teacher/w3/grade_list.xlsx'


# Load the Excel workbook and select the sheet
workbook = openpyxl.load_workbook(excel_file_path)
# Get the sheet names
sheet_names = workbook.sheetnames
sheet_name=sheet_names[0]
# Print the sheet names
#print("Sheet names:", sheet_names)
sheet = workbook[sheet_name]

# Specify the cell coordinates (row and column indices, 1-based index)
# Access the cell range
cell_range_pattern='c5:j12'
cell_range = sheet[cell_range_pattern]
grade_list=[]

for row in sheet['c5:j12']:
   row_value=[]
   for cell in row:
        row_value.append(cell.value)
   grade_list.append(row_value)

#POP
#total and average score
for i in range(1,len(grade_list)): 
    total=0
    student=grade_list[i]
    for j in range(1,5):
        total+=student[j]
    average=total/4
    student[5]=total
    student[6]=average

#rank
for i in range(1,8):
    k=1
    for j in range(1,8):
        if(grade_list[i][5] < grade_list[j][5]):
            if(i!=j):
                k=k+1
    grade_list[i][7]=k
    
#print
print('姓名\t國文\t英文\t數學\t理化\t總分\t平均\t名次')
print('------------------------------------------------------------')
for i in range(len(grade_list)):
    print(grade_list[i][0], end='\t')
    for j in range(1, len(grade_list[i])):
        print(grade_list[i][j], end='\t')
    print()

#OOP
class Grade_list():
    def __init__(file, path):
        file.path=path
        file.workbook=openpyxl.load_workbook(path)
        file.cell_range_pattern='c5:j12'
        file.grade_list=[]
         
        file.get_data()
        file.total_score()
        file.average()
        file.rank()
        file.print()
        file.close()

    def get_data(file):
        file.sheet_names=file.workbook.sheetnames
        file.sheet_name=file.sheet_names[0]
        sheet_name=file.sheet_names[0]
        sheet=file.workbook[sheet_name]

        for row in sheet[cell_range_pattern]:
            row_values=[]
            for cell in row:
                row_values.append(cell.value)
            file.grade_list.append(row_values)

    def total_score(file):
        for i in range(1,len(file.grade_list)): 
            total=0
            student=file.grade_list[i]
            for j in range(1,5):
                total+=student[j]
                file.grade_list[i][5]=total

    def average(file):
        for i in range(1,len(file.grade_list)):
            file.grade_list[i][6]=file.grade_list[i][5]/4
    
    def rank(file):
        for i in range(1,8):
            k=1
            for j in range(1,8):
                if(file.grade_list[i][5] < file.grade_list[j][5]):
                    if(i!=j):
                        k=k+1
            file.grade_list[i][7]=k
        
    def print(file):
        print('姓名\t國文\t英文\t數學\t理化\t總分\t平均\t名次')
        print('------------------------------------------------------------')
        for i in range(len(file.grade_list)):
            print(file.grade_list[i][0], end='\t')
            for j in range(1, len(file.grade_list[i])):
                print(file.grade_list[i][j], end='\t')
            print()
        print()

    def close(file):
        file.workbook.close()

w3_hw = Grade_list('D:/MyData/Desktop/PYTHON/python_course/teacher/w3/grade_list.xlsx')