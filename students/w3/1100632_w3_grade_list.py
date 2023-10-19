

import openpyxl

print("pop")
def getdata(excel_file_path):
    # Load the Excel workbook and select the sheet
    workbook = openpyxl.load_workbook(excel_file_path)
    # Get the sheet names
    sheet_names = workbook.sheetnames
    sheet_name=sheet_names[0]
    # Print the sheet names
    
    sheet = workbook[sheet_name]


    cell_range_pattern='c5:j12'
    cell_range = sheet[cell_range_pattern]

    grade_list=[]
    for row in cell_range:
        rowlist=[]
        for cell in row:
            rowlist.append(cell.value)
        grade_list.append(rowlist)
    workbook.close()    
    return grade_list



def sum_average_and_sorted(grade_list):
    
    #計算平均、總和
    for i in range(1,len(grade_list)):
        grade_list[i][5]=sum(grade_list[i][1:5])
        grade_list[i][6]=grade_list[i][5]/4
    #排列名次
    grade_list_sort = sorted(grade_list[1:len(grade_list)], key=lambda x: x[5], reverse=True)
    #將名次放進去
    for i in range(len(grade_list_sort)):
            grade_list_sort[i][7]=i+1
    return grade_list_sort



def print_grade_list(grade_list):
    print('姓名\t國文\t英文\t數學\t理化\t總分\t平均\t名次')
    print('------------------------------------------------------------')
    for i in range(len(grade_list)):
        print(grade_list[i][0], end='\t')
        for j in range(1, len(grade_list[i])):
            print(grade_list[i][j], end='\t')
        print()


print_grade_list(sum_average_and_sorted(getdata(r'C:\Users\kmsh1\Desktop\python程式設計\python_course\teacher\w3\grade_list.xlsx')))


print("\n")
print("oop")


class gradelist:

    def __init__(self,excel_file_path):
        self.excel_file_path=excel_file_path
    
    def getdata(self):
    # Load the Excel workbook and select the sheet
        workbook = openpyxl.load_workbook(self.excel_file_path)
        # Get the sheet names
        sheet_names = workbook.sheetnames
        sheet_name=sheet_names[0]
        # Print the sheet names
        
        sheet = workbook[sheet_name]


        cell_range_pattern='c5:j12'
        cell_range = sheet[cell_range_pattern]

        self.grade_list=[]
        for row in cell_range:
            rowlist=[]
            for cell in row:
                rowlist.append(cell.value)
            self.grade_list.append(rowlist)
        workbook.close()    
        return self.grade_list



    def sum_average_and_sorted(self):
        
        #計算平均、總和
        for i in range(1,len(self.grade_list)):
            self.grade_list[i][5]=sum(self.grade_list[i][1:5])
            self.grade_list[i][6]=self.grade_list[i][5]/4
        #排列名次
        self.grade_list_sort = sorted(self.grade_list[1:len(self.grade_list)], key=lambda x: x[5], reverse=True)
        #將名次放進去
        for i in range(len(self.grade_list_sort)):
                self.grade_list_sort[i][7]=i+1
        return self.grade_list_sort



    def print_grade_list(self):
        print('姓名\t國文\t英文\t數學\t理化\t總分\t平均\t名次')
        print('------------------------------------------------------------')
        for i in range(len(self.grade_list)):
            print(self.grade_list[i][0], end='\t')
            for j in range(1, len(self.grade_list[i])):
                print(self.grade_list[i][j], end='\t')
            print()


grade=gradelist(r'C:\Users\kmsh1\Desktop\python程式設計\python_course\teacher\w3\grade_list.xlsx')
grade.getdata()
grade.sum_average_and_sorted()
grade.print_grade_list()

