import openpyxl

print("POP")
excel_file_path = "C:/Users/user/Desktop/shell/pythonCourse/dataset/grade_list.xlsx"
workbook = openpyxl.load_workbook(excel_file_path)
sheet_names = workbook.sheetnames
sheet = workbook[sheet_names[0]]
cell_range = sheet['c5:j12']
grade_list = [[data.value for data in row] for row in cell_range]
workbook.close()
rankList = []

for i in range(1, 8):
    grade_list[i][5] = 0
    for j in range(1, 5):
        grade_list[i][5] += grade_list[i][j]
    rankList.append(grade_list[i][5])
    grade_list[i][6] = 0
    grade_list[i][6] = grade_list[i][5] / 4

rankList.sort()
for i in range(1, 8):
    grade_list[i][7] = rankList.index(grade_list[i][5]) + 1

for i in range(8):
    for j in range(8):
        print(grade_list[i][j], end='\t')
    if i == 0:
        print('\n------------------------------------------------------------', end="")
    print()

print("OOP")


class Grade:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path

    def accessToData(self, dataRange):
        workbook = openpyxl.load_workbook(self.excel_file_path)
        sheet_names = workbook.sheetnames
        sheet = workbook[sheet_names[0]]
        cell_range = sheet[dataRange]
        grade_list = [[data.value for data in row] for row in cell_range]
        workbook.close()
        return grade_list

    def calSumAndAvg(self, grade_list):
        rankList = []
        for i in range(1, 8):
            grade_list[i][5] = 0
            for j in range(1, 5):
                grade_list[i][5] += grade_list[i][j]
            rankList.append(grade_list[i][5])
            grade_list[i][6] = 0
            grade_list[i][6] = grade_list[i][5] / 4
        return grade_list, rankList

    def calSort(self, grade_list, rankList):
        rankList.sort()
        for i in range(1, 8):
            grade_list[i][7] = rankList.index(grade_list[i][5]) + 1
        return grade_list

    def printing(self, grade_list):
        for i in range(8):
            for j in range(8):
                print(grade_list[i][j], end='\t')
            if i == 0:
                print(
                    '\n------------------------------------------------------------', end="")
            print()


grade = Grade(
    "C:/Users/user/Desktop/shell/pythonCourse/dataset/grade_list.xlsx")
grade_list = grade.accessToData('c5:j12')
grade_list, rankList = grade.calSumAndAvg(grade_list)
grade_list = grade.calSort(grade_list, rankList)
grade.printing(grade_list)
